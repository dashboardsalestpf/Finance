import streamlit as st
st.set_page_config(layout="wide")

import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import altair as alt
import json
import datetime
import time
from google.oauth2.service_account import Credentials
import gspread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io


@st.cache_resource
def connect_gsheet(x):
    creds_dict = st.secrets["projectfinance"]
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    sheet = client.open_by_key("1TbrJybGgTjJowu2sfFw4s3NFcAD580y4O8MvEEZtWuw").worksheet(x)
    return sheet

# ----- Load Google Sheet into DataFrame -----
def load_data(x):
    sheet = connect_gsheet(x)
    data = sheet.get_all_records()
    df = pd.DataFrame(data)
    return df

@st.cache_data
def loading_data():
    df = load_data("Sheet1")
    df2 = load_data("m_branch")
    df3 = load_data("m_source")
    df4 = load_data("m_vendor")
    return df, df2, df3, df4

def export_excel(df, branch_name, payment_date, selected_vendor):

    wb = Workbook()
    ws = wb.active

    # ===== Styles =====
    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    left = Alignment(horizontal='left')
    right = Alignment(horizontal='right')

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Header =====
    ws['A1'] = branch_name
    ws['A1'].font = Font(bold=True, underline='single', size=12)
    ws['A1'].alignment = left

    ws.merge_cells('B1:D1')
    ws['B1'] = "FORM PENGAJUAN PEMBAYARAN"
    ws['B1'].font = Font(bold=True, underline='single', size=14)
    ws['B1'].alignment = left

    ws['A3'] = f"Vendor : {str(selected_vendor)}"
    ws['A3'].font = Font(bold=True, underline='single', size=12)
    ws['A3'].alignment = left

    ws['D3'] = f"Tgl Terima : {str(payment_date)}"
    ws['D3'].font = Font(size=12)
    ws['D3'].alignment = right

    ws['D4'] = f"Tgl Bayar : {str(payment_date)}"
    ws['D4'].font = Font(size=12)
    ws['D4'].alignment = right

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    # ===== Table Header =====
    headers = ["No Invoice", "Keterangan", "Total"]

    start_row = 6

    ws.cell(row=start_row, column=1, value=headers[0]).font = bold
    ws.cell(row=start_row, column=1).alignment = left

    ws.merge_cells('B6:C6')
    ws.cell(row=start_row, column=2, value=headers[1]).font = bold
    ws.cell(row=start_row, column=2).alignment = left

    ws.cell(row=start_row, column=4, value=headers[2]).font = bold
    ws.cell(row=start_row, column=4).alignment = left

    # ===== Data =====
    row_num = start_row + 1

    for _, row in df.iterrows():
        # merge B and C for THIS row
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)

        # write only into top-left cell (B)
        ws.cell(row=row_num, column=1, value=row['No Invoice'])
        ws.cell(row=row_num, column=2, value=row['Keterangan'])
        ws.cell(row=row_num, column=4, value=row['Nominal'])

        row_num += 1

    # ===== Grand Total =====
    # ws.cell(row=row_num, column=1, value="Grand Total").font = Font(bold=True, size=14)
    # ws.cell(row=row_num, column=4, value=grand_total).font = Font(bold=True, size=14)

    # ===== Data rows end position =====
    last_data_row = row_num - 1

    # ===== Apply border ONLY for data + grand total =====
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=last_data_row,
        min_col=1,
        max_col=4
    ):
        for cell in row:
            cell.border = thin

    # ===== Make last row (Grand Total) bold + size 14 =====
    for col in range(1, 5):
        ws.cell(row=last_data_row, column=col).font = Font(bold=True, size=14)

    # ===== Number format for all data rows =====
    for row in range(start_row + 1, last_data_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0.00'

    # Grand total row formatting (redundant but safe)
    ws.cell(row=last_data_row, column=4).number_format = '#,##0.00'

    # ===== Footer =====
    row_num += 3
    ws.cell(row=row_num, column=1, value="Diajukan,")
    ws.cell(row=row_num, column=2, value="Disetujui,")
    ws.cell(row=row_num, column=4, value="Mengetahui,")

    # ===== Save to memory =====
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

df, df2, df3, df4 = loading_data()

st.title("Pivot")

# Branch filter
name_to_branch = dict(zip(df2['Name'], df2['Branch']))

selected_name = st.selectbox(
    "Branch",
    df2['Name'].sort_values().unique()
)

Branch = name_to_branch[selected_name]

# Date filter
Date2 = st.date_input("Tanggal Payment")

# Vendor filter
vendor_list = df['Vendor'].dropna().unique()
selected_vendor = st.selectbox("Vendor", sorted(vendor_list))

# Convert date column to datetime (IMPORTANT)
df['Payment Date'] = pd.to_datetime(df['Payment Date'], dayfirst=True)

# Apply filters
filtered_df = df[
    (df['Branch'] == Branch) &
    (df['Payment Date'].dt.date == Date2) &
    (df['Vendor'] == selected_vendor)
]

pivot_df = filtered_df[['No Invoice', 'Keterangan', 'Nominal']].copy()

pivot_df = pivot_df.sort_values(by=['No Invoice'])

dup_invoice = pivot_df.duplicated(subset=['No Invoice'])

pivot_df.loc[dup_invoice, 'No Invoice'] = ''

grand_total = pivot_df['Nominal'].sum()

total_row = pd.DataFrame({
    'No Invoice': [''],
    'Keterangan': ['Grand Total'],
    'Nominal': [grand_total]
})

pivot_df = pd.concat([pivot_df, total_row], ignore_index=True)

st.dataframe(pivot_df, use_container_width=True, hide_index=True)

excel_file = export_excel(
    pivot_df,
    selected_name,
    Date2,
    selected_vendor
)

st.download_button(
    label="⬇️ Download Excel Report",
    data=excel_file,
    file_name="payment_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)