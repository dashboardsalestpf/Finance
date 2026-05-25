import streamlit as st
st.set_page_config(layout="wide")

import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
import altair as alt
import json
import datetime
import time
from google.oauth2.service_account import Credentials
import gspread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io


@st.cache_resource
def connect_gsheet(x):
    creds_dict = st.secrets["projectfinance"]
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    sheet = client.open_by_key("1TbrJybGgTjJowu2sfFw4s3NFcAD580y4O8MvEEZtWuw").worksheet(x)
    return sheet

# ----- Load Google Sheet into DataFrame -----
def load_data(x):
    sheet = connect_gsheet(x)
    data = sheet.get_all_records()
    df = pd.DataFrame(data)
    return df

@st.cache_data
def loading_data():
    df = load_data("Sheet1")
    df2 = load_data("m_branch")
    df3 = load_data("m_source")
    df4 = load_data("m_vendor")
    return df, df2, df3, df4

def export_excel(df, branch_name, payment_date, selected_vendor):

    wb = Workbook()
    ws = wb.active

    # ===== Styles =====
    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    left = Alignment(horizontal='left')
    right = Alignment(horizontal='right')

    # ===== Header =====
    ws['A1'] = branch_name
    ws['A1'].font = Font(bold=True, underline='single', size=12)

    ws.merge_cells('A2:C2')
    ws['A2'] = "FORM PENGAJUAN PEMBAYARAN"
    ws['A2'].font = Font(bold=True, underline='single', size=14)
    ws['A2'].alignment = center

    ws['C3'] = f"Tgl Terima : {str(payment_date)}"
    ws['C3'].alignment = right

    ws['C4'] = f"Tgl Bayar : {str(payment_date)}"
    ws['C4'].alignment = right

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30

    # ===== Prepare List Data =====
    clean_df = df[df['Keterangan'] != 'Grand Total'].copy()

    invoice_list = ", ".join(clean_df['No Invoice'].replace('', pd.NA).dropna())
    keterangan_list = ", ".join(clean_df['Keterangan'].dropna())
    grand_total = clean_df['Nominal'].sum()

    # ===== Write List Layout =====
    start_row = 7

    ws.cell(row=start_row-1, column=1, value="Vendor").font = bold
    ws.cell(row=start_row-1, column=2, value=f" : {selected_vendor}")

    ws.cell(row=start_row, column=1, value="No Invoice").font = bold
    ws.cell(row=start_row, column=2, value=f" : {invoice_list}")

    ws.cell(row=start_row + 1, column=1, value="Keterangan").font = bold
    ws.cell(row=start_row + 1, column=2, value=f" : {keterangan_list}")

    ws.cell(row=start_row + 2, column=1, value="Total").font = bold
    ws.cell(row=start_row + 2, column=2, value=f" : {grand_total}")

    ws.cell(row=start_row + 2, column=2).alignment = Alignment(horizontal='left')
    ws.cell(row=start_row + 2, column=2).number_format = '#,##0.00'

    # Wrap text (important for long lists)
    ws.cell(row=start_row, column=2).alignment = Alignment(wrap_text=True)
    ws.cell(row=start_row + 1, column=2).alignment = Alignment(wrap_text=True)

    # ===== Footer =====
    footer_row = start_row + 6

    ws.cell(row=footer_row, column=1, value="Diajukan,")
    ws.cell(row=footer_row, column=2, value="Disetujui,")
    ws.cell(row=footer_row, column=3, value="Mengetahui,")

    # ===== Save =====
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

df, df2, df3, df4 = loading_data()

st.title("Pivot")

# Branch filter
name_to_branch = dict(zip(df2['Name'], df2['Branch']))

selected_name = st.selectbox(
    "Branch",
    df2['Name'].sort_values().unique()
)

Branch = name_to_branch[selected_name]

# Date filter
Date2 = st.date_input("Tanggal Payment")

# Vendor filter
vendor_list = df['Vendor'].dropna().unique()
selected_vendor = st.selectbox("Vendor", sorted(vendor_list))

# Convert date column to datetime (IMPORTANT)
df['Payment Date'] = pd.to_datetime(df['Payment Date'], dayfirst=True)

# Apply filters
filtered_df = df[
    (df['Branch'] == Branch) &
    (df['Payment Date'].dt.date == Date2) &
    (df['Vendor'] == selected_vendor)
]

pivot_df = filtered_df[['No Invoice', 'Keterangan', 'Nominal']].copy()

pivot_df = pivot_df.sort_values(by=['No Invoice'])

dup_invoice = pivot_df.duplicated(subset=['No Invoice'])

pivot_df.loc[dup_invoice, 'No Invoice'] = ''

grand_total = pivot_df['Nominal'].sum()

total_row = pd.DataFrame({
    'No Invoice': [''],
    'Keterangan': ['Grand Total'],
    'Nominal': [grand_total]
})

pivot_df = pd.concat([pivot_df, total_row], ignore_index=True)

st.dataframe(pivot_df, use_container_width=True, hide_index=True)

excel_file = export_excel(
    pivot_df,
    selected_name,
    Date2,
    selected_vendor
)

st.download_button(
    label="⬇️ Download Excel Report",
    data=excel_file,
    file_name="payment_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)